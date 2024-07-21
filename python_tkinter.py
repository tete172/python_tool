import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import pandas as pd
from tkinter import filedialog

def load_csv(label):
    global file_path_1
    # csvファイル選択ダイアログを開く
    file_path_1 = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    
    if file_path_1:
        # 取得したファイルパスをラベルに表示
        file_name_1 = os.path.basename(file_path_1)
        label.config(text=file_name_1)
        
def load_xlsx(label,combobox):
    global file_path_2
    # xlsxファイル選択ダイアログを開く
    file_path_2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if file_path_2:
        # 取得したファイルパスをラベルに表示
        file_name_2 = os.path.basename(file_path_2)
        
        label.config(text=file_name_2)
        
        workbook = load_workbook(file_path_2)
        # シート名のリストを取得する
        sheet_names = workbook.sheetnames
        # プルダウンBにシート名を反映
        combobox['values'] = sheet_names
        if sheet_names:  # シート名が存在する場合は、最初のシート名を選択
            combobox.set(sheet_names[0])
        
def load_folder(label):
    global file_path_3
    # フォルダ選択ダイアログを開く
    file_path_3 = filedialog.askdirectory()
    file_name_3 = os.path.basename(file_path_3)
    
    label.config(text=file_name_3)
    
    
            
def execution_button_1():
    sheat_name = sheat_combo_1.get()
    csv_name = loadcsv_label_1.cget("text")
    xlsx_name = loadxlsx_label_1.cget("text")

    # CSVファイルを読み込む
    csv_data = pd.read_csv(file_path_1, encoding="shift_jis")

    # Excelファイルを読み込む
    wb = wb = load_workbook(file_path_2, data_only=False)
    ws = wb[sheat_name]

    # ヘッダーを含むデータフレームをExcelシートに貼り付ける
    rows = dataframe_to_rows(csv_data, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 各シートの再計算を実行
    for sheet in wb.sheetnames:
        wb[sheet].calculate_dimension()
        time.sleep(1)
    
    print(f'{csv_name} の内容を {xlsx_name} のシート名：{sheat_name}に貼り付けました')
    wb.save(file_path_2)

    wb.close()
    
    
def execution_button_2():
    sheat_name_2 = sheat_combo_2.get()

    filename_2 = filename_entry_2.get() + ".csv"
    
    output_csv = os.path.join(file_path_3,filename_2)
    
    df = pd.read_excel(file_path_2, sheet_name=sheet_name_2)  

    # CSVとして保存
    df.to_csv(output_csv, encoding="shift_jis", index=False)
    
    print(f'{filename_2} シート名：{sheat_name_2}の内容を {output_csv}に保存しました')
    
    
def execution_button_3():
    input_csv_3 = file_path_1
    path_name_3 = filename_entry_3.get() + "_{}.csv"
    section_name = os.path.join(file_path_3, path_name_3)
    chunk_size = int(chunk_entry_3.get())
    
    # x行づつCSVファイルを分割する-----ここから
    def split_and_save_to_excel(input_csv_3, chunk_size):
        # CSVファイルを読み取り、DataFrameに格納する
        chunks = pd.read_csv(input_csv_3, chunksize=chunk_size, header=None, encoding="shift_jis")

        # ヘッダーを取得
        header = pd.read_csv(input_csv_3, nrows=1, header=None, encoding="shift_jis").iloc[0]

        # 分割したデータを新しいCSVファイルとして保存する
        for i, chunk in enumerate(chunks):
            file_section = section_name.format(i+1)  # 新しいCSVファイルの名前
            # ヘッダー行を先頭に追加して保存
            if i == 0:
                chunk = chunk.iloc[1:]
            chunk.to_csv(file_section, index=False, header=False, mode='a', encoding="shift_jis")

            # 先頭にヘッダー行を挿入
            with open(file_section, 'r+', encoding="shift_jis") as f:
                content = f.read()
                f.seek(0, 0)
                f.write(','.join(map(str, header)) + '\n' + content)

    # 分割してCSVファイルとして保存
    split_and_save_to_excel(input_csv_3, chunk_size)

    # x行づつCSVファイルを分割する-----ここまで
    print(f"{input_csv_3}を分割して保存しました")
    
    


# メインウィンドウの作成
root = tk.Tk()
root.title("CSV・xlsxツール")
root.geometry("550x350")

# フレーム1の作成
frame1 = tk.Frame(root)
label1 = tk.Label(frame1, text="xlsxにCSVを貼り付け", padx=5, pady=5, width=20, height=1)
label1.grid(row=0, column=0, padx=5, pady=5)

button1 = tk.Button(frame1, text="xlsxにCSVを貼り付け", command=lambda: show_frame(frame1), width=20, height=1)
button1.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)

button2 = tk.Button(frame1, text="xlsxを読込みCSV保存", command=lambda: show_frame(frame2), width=20, height=1)
button2.grid(row=1, column=1, sticky='nsew', padx=5, pady=5)

button3 = tk.Button(frame1, text="CSVを読込み分割", command=lambda: show_frame(frame3), width=20, height=1)
button3.grid(row=1, column=2, sticky='nsew', padx=5, pady=5)

# フレーム1 1列名-------------------------------------------------------------------------ここから
# CSVを読み込むボタン
load_csv_1 = tk.Button(frame1, text="csvを読み込む", command=lambda: load_csv(loadcsv_label_1),
                       width=20, height=1,
                       borderwidth=1,
                       relief="raised",
                       padx=3, pady=3,
                       font=("Arial", 10, "normal"),
                       fg="black", bg="white")

# ボタンload_csvをウィンドウに配置
load_csv_1.grid(row=2, column=0, padx=5, pady=5)
# xlsxを読み込むボタン
load_xlsx_1 = tk.Button(frame1,text="貼付け先xlsx", command=lambda: load_xlsx(loadxlsx_label_1,sheat_combo_1),
                        width=20, height=1,
                        borderwidth=1,      # 枠の幅を指定
                        relief="raised",    # 枠のスタイルを指定
                        padx=3, pady=3,   # パディングを指定
                        font=("Arial", 10, "normal"), # フォントを指定
                        fg="black", bg="white")  # 文字色と背景色を指定)
                        
# ボタンload_xlsxをウィンドウに配置
load_xlsx_1.grid(row=3, column=0, padx=5, pady=5)

# ラベル貼付け先シート名
sheat_label_1 = tk.Label(frame1,text="貼付け先シート名",
                         width=20, height=1,
                         borderwidth=1,      # 枠の幅を指定
                         relief="raised",    # 枠のスタイルを指定
                         padx=3, pady=3,   # パディングを指定
                         font=("Arial", 10, "normal"), # フォントを指定
                         fg="black", bg="white")         # 文字色と背景色を指定)
                 
# ボタンload_xlsxをウィンドウに配置
sheat_label_1.grid(row=4, column=0, padx=5, pady=5)

# フレーム1 1列名-------------------------------------------------------------------------ここまで
# フレーム1 2列目-------------------------------------------------------------------------ここから

loadcsv_label_1 = tk.Label(frame1, text="No file selected",
                           width=20, height=1,
                           borderwidth=0,
                           relief="raised",
                           padx=3, pady=3,
                           font=("Arial", 10, "normal"),
                           fg="black", bg="white")
                     
loadcsv_label_1.grid(row=2, column=1, padx=5, pady=5)

# 2列4行目
loadxlsx_label_1 = tk.Label(frame1, text="No file selected",
                            width=20, height=1,
                            borderwidth=0,
                            relief="raised",
                            padx=3, pady=3,
                            font=("Arial", 10, "normal"),
                            fg="black", bg="white")
                     
loadxlsx_label_1.grid(row=3, column=1, padx=5, pady=5)

# 2列5行目
# プルダウンシート名表示
sheat_combo_1 = ttk.Combobox(frame1, width=20)
sheat_combo_1 .grid(row=4, column=1, padx=5, pady=5)

# 2列6行目
execution_button_1 = tk.Button(frame1,text="実行する", command=execution_button_1,
                               width=20, height=1,
                               borderwidth=1,      # 枠の幅を指定
                               relief="raised",    # 枠のスタイルを指定
                               padx=3, pady=3,   # パディングを指定
                               font=("Arial", 10, "normal"), # フォントを指定
                               fg="black", bg="white")  # 文字色と背景色を指定)
                 
# ラベルbをウィンドウに配置
execution_button_1.grid(row=5, column=1, padx=5, pady=5)
# フレーム1 2列目-------------------------------------------------------------------------ここまで

# フレーム2の作成
frame2 = tk.Frame(root)
label2 = tk.Label(frame2, text="xlsxを読込みCSV保存", padx=5, pady=5)
label2.grid(row=0, column=0, padx=5, pady=5)

button1 = tk.Button(frame2, text="xlsxにCSVを貼り付け", command=lambda: show_frame(frame1), width=20, height=1)
button1.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)

button2 = tk.Button(frame2, text="xlsxを読込みCSV保存", command=lambda: show_frame(frame2), width=20, height=1)
button2.grid(row=1, column=1, sticky='nsew', padx=5, pady=5)

button3 = tk.Button(frame2, text="CSVを読込み分割", command=lambda: show_frame(frame3), width=20, height=1)
button3.grid(row=1, column=2, sticky='nsew', padx=5, pady=5)

# フレーム2 1列目-------------------------------------------------------------------------ここから
# ボタンxlsxを読み込む
load_xlsx_2 = tk.Button(frame2, text="xlsxを読み込む", command=lambda: load_xlsx(loadxlsx_label_2,sheat_combo_2 ),
                        width=20, height=1,
                        borderwidth=1,
                        relief="raised",
                        padx=3, pady=3,
                        font=("Arial", 10, "normal"),
                        fg="black", bg="white")
                    
# ボタンload_csvをウィンドウに配置
load_xlsx_2.grid(row=2, column=0, padx=5, pady=5)

# ラベル貼付け先シート名
sheat_label_2 = tk.Label(frame2,text="貼付け先シート名",
                         width=20, height=1,
                         borderwidth=1,      # 枠の幅を指定
                         relief="raised",    # 枠のスタイルを指定
                         padx=3, pady=3,   # パディングを指定
                         font=("Arial", 10, "normal"), # フォントを指定
                         fg="black", bg="white")  # 文字色と背景色を指定)
                 
# sheat_label_2をウィンドウに配置
sheat_label_2.grid(row=3, column=0, padx=5, pady=5)

# ボタンフォルダを指定する
load_folder_2 = tk.Button(frame2, text="保存フォルダを指定", command=lambda: load_folder(folder_label_2),
                          width=20, height=1,
                          borderwidth=1,
                          relief="raised",
                          padx=3, pady=3,
                          font=("Arial", 10, "normal"),
                          fg="black", bg="white")
                    
# load_folder_2をウィンドウに配置
load_folder_2.grid(row=4, column=0, padx=5, pady=5)

# ラベル保存先ファイル名
filename_label_2 = tk.Label(frame2,text="保存先ファイル名",
                            width=20, height=1,
                            borderwidth=1,      # 枠の幅を指定
                            relief="raised",    # 枠のスタイルを指定
                            padx=3, pady=3,   # パディングを指定
                            font=("Arial", 10, "normal"), # フォントを指定
                            fg="black", bg="white")  # 文字色と背景色を指定)
                 
# filename_label_2をウィンドウに配置
filename_label_2.grid(row=5, column=0, padx=5, pady=5)
# フレーム2 1列目-------------------------------------------------------------------------ここまで
# フレーム2 2列目-------------------------------------------------------------------------ここから

#フレーム2 2列目3行目から
loadxlsx_label_2 = tk.Label(frame2, text="No file selected",
                            width=20, height=1,
                            borderwidth=0,
                            relief="raised",
                            padx=3, pady=3,
                            font=("Arial", 10, "normal"),
                            fg="black", bg="white")
                     
loadxlsx_label_2.grid(row=2, column=1, padx=5, pady=5)

#フレーム2 2列目4行目
# プルダウンシート名表示
sheat_combo_2 = ttk.Combobox(frame2, width=20)
sheat_combo_2 .grid(row=3, column=1, padx=5, pady=5)

#フレーム2 2列目5行目
folder_label_2 = tk.Label(frame2, text="No file selected",
                          width=20, height=1,
                          borderwidth=0,
                          relief="raised",
                          padx=3, pady=3,
                          font=("Arial", 10, "normal"),
                          fg="black", bg="white")
                     
folder_label_2.grid(row=4, column=1, padx=5, pady=5)

#フレーム2 2列目5行目
filename_entry_2 = tk.Entry(frame2, width=30)
filename_entry_2.grid(row=5, column=1, padx=5, pady=5)


# フレーム2列6行目
execution_button_2 = tk.Button(frame2, text="実行する", command=execution_button_2,
                               width=20, height=1,
                               borderwidth=1,      # 枠の幅を指定
                               relief="raised",    # 枠のスタイルを指定
                               padx=3, pady=3,   # パディングを指定
                               font=("Arial", 10, "normal"), # フォントを指定
                               fg="black", bg="white")  # 文字色と背景色を指定)
                 
# ラベルbをウィンドウに配置
execution_button_2.grid(row=6, column=1, padx=5, pady=5)

# フレーム2 2列目-------------------------------------------------------------------------ここまで
# フレーム3の作成
frame3 = tk.Frame(root)
label3 = tk.Label(frame3, text="CSVファイルを読み込み分割", padx=5, pady=5)
label3.grid(row=0, column=0, padx=5, pady=5)

button1 = tk.Button(frame3, text="xlsxにCSVを貼り付け", command=lambda: show_frame(frame1), width=20, height=1)
button1.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)

button2 = tk.Button(frame3, text="xlsxを読込みCSV保存", command=lambda: show_frame(frame2), width=20, height=1)
button2.grid(row=1, column=1, sticky='nsew', padx=5, pady=5)

button3 = tk.Button(frame3, text="CSVを読込み分割", command=lambda: show_frame(frame3), width=20, height=1)
button3.grid(row=1, column=2, sticky='nsew', padx=5, pady=5)
# フレーム3 1列目-------------------------------------------------------------------------ここから
# CSVを読み込むボタン
load_csv_3 = tk.Button(frame3, text="csvを読み込む", command=lambda: load_csv(loadcsv_label_3),
                       width=20, height=1,
                       borderwidth=1,
                       relief="raised",
                       padx=3, pady=3,
                       font=("Arial", 10, "normal"),
                       fg="black", bg="white")
                    
# ボタンload_csvをウィンドウに配置
load_csv_3.grid(row=2, column=0, padx=5, pady=5)

# フォルダを指定するボタン
load_folder_3 = tk.Button(frame3, text="保存フォルダを指定", command=lambda: load_folder(folder_label_3),
                          width=20, height=1,
                          borderwidth=1,
                          relief="raised",
                          padx=3, pady=3,
                          font=("Arial", 10, "normal"),
                          fg="black", bg="white")
                    
# フォルダを指定するボタンをウィンドウに配置
load_folder_3.grid(row=3, column=0, padx=5, pady=5)

# 保存先ファイル名
filename_label_3 = tk.Label(frame3,text="保存先ファイル名",
                            width=20, height=1,
                            borderwidth=1,      # 枠の幅を指定
                            relief="raised",    # 枠のスタイルを指定
                            padx=3, pady=3,   # パディングを指定
                            font=("Arial", 10, "normal"), # フォントを指定
                            fg="black", bg="white")         # 文字色と背景色を指定)
                 
# 保存先ファイル名をウィンドウに配置
filename_label_3.grid(row=4, column=0, padx=5, pady=5)

# チャンクサイズ指定
chunk_size_3 = tk.Label(frame3,text="分割数",
                        width=20, height=1,
                        borderwidth=1,      # 枠の幅を指定
                        relief="raised",    # 枠のスタイルを指定
                        padx=3, pady=3,   # パディングを指定
                        font=("Arial", 10, "normal"), # フォントを指定
                        fg="black", bg="white")         # 文字色と背景色を指定)
                 
# 保存先ファイル名をウィンドウに配置
chunk_size_3.grid(row=5, column=0, padx=5, pady=5)

# フレーム3 1列目-------------------------------------------------------------------------ここまで

# フレーム3 2列目-------------------------------------------------------------------------ここから
# フレーム3 2列目3行目
loadcsv_label_3 = tk.Label(frame3, text="No file selected",
                           width=20, height=1,
                           borderwidth=0,
                           relief="raised",
                           padx=3, pady=3,
                           font=("Arial", 10, "normal"),
                           fg="black", bg="white")
                     
loadcsv_label_3.grid(row=2, column=1, padx=5, pady=5)

# フレーム3 2列目4行目
folder_label_3 = tk.Label(frame3, text="No file selected",
                          width=20, height=1,
                          borderwidth=0,
                          relief="raised",
                          padx=3, pady=3,
                          font=("Arial", 10, "normal"),
                          fg="black", bg="white")
                     
folder_label_3.grid(row=3, column=1, padx=5, pady=5)

# フレーム3 2列目5行目
filename_entry_3 = tk.Entry(frame3, width=30)
filename_entry_3.grid(row=4, column=1, padx=5, pady=5)

# フレーム3 2列目6行目
chunk_entry_3 = tk.Entry(frame3, width=30)
chunk_entry_3 .grid(row=5, column=1, padx=5, pady=5)

# フレーム3列7行目
execution_button_3 = tk.Button(frame3, text="実行する", command=execution_button_3,
                               width=20, height=1,
                               borderwidth=1,      # 枠の幅を指定
                               relief="raised",    # 枠のスタイルを指定
                               padx=3, pady=3,   # パディングを指定
                               font=("Arial", 10, "normal"), # フォントを指定
                               fg="black", bg="white")  # 文字色と背景色を指定)
                 
# ラベルbをウィンドウに配置
execution_button_3.grid(row=6, column=1, padx=5, pady=5)

# フレーム3 2列目-------------------------------------------------------------------------ここまで



# フレームを表示する関数
def show_frame(frame):
    frame.tkraise()

# 最初にフレーム1を表示
for frame in (frame1, frame2, frame3):
    frame.grid(row=0, column=0, sticky='nsew')

show_frame(frame1)



# メインループの開始
root.mainloop()
